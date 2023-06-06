import openpyxl
import json

our_book_path = "miranda/Таблица БС.xlsx"
current_alarms_path = "miranda/CurrentAlarmsTest.xlsx"
ne_list_path = "miranda/NE Report_test.xlsx"
all_addresses_path = "miranda/Адресный список  всех БС.xlsx"



# Читаем наши станции
def read_our_book(path):
    book = openpyxl.load_workbook(filename=path)
    our_bs_list = []
    counter = 0

    for sheet in book.worksheets[1:]:
        for i in range(1,61):
            counter += 1
            if sheet[f'b{i}'].value != None:
                our_bs_list.append(sheet[f'b{i}'].value[:2].strip() + sheet[f'b{i}'].value[-4:].strip())

    return our_bs_list


# Читаем NE_list
def read_ne_list(path):
    book = openpyxl.load_workbook(filename=path)
    ne_list = []
    counter = 0

    for sheet in book.worksheets:
        for i in range(1, 1000):
            counter += 1
            if sheet[f'a{i}'].value != None:
                if sheet[f'q{i}'].value == "Offline":

                    ne_list.append(sheet[f'a{i}'].value[:2] + sheet[f'a{i}'].value[-4:])
    return ne_list


# Читаем Current_alarms
def read_current_alarms(path):
    book = openpyxl.load_workbook(filename=path)
    cur_al_list = []
    counter = 0

    for sheet in book.worksheets:

        for i in range(1, 4000):

            if sheet[f's{i}'].value != None:
                if sheet[f's{i}'].value != "-":
                    if sheet[f's{i}'].value != "BBU Name":
                        counter += 1

                        cur_al_list.append(sheet[f's{i}'].value)
    # print("Все алармы :",cur_al_list)
    # print(len(cur_al_list))
    return cur_al_list


# Читаем все адреса станций
def read_all_addresses(path):
    pass

# Ищем наши NE станции
def common_ne(list_1,list_2):
    common_ne_list = list(set(list_1) & set(list_2))
    print("наши БС не в сети:", common_ne_list)
    print("количество наших БС не в сети: ",len(common_ne_list))
    return common_ne_list


# Ищем наши current alams
def common_current_alarms(list_1,list_2):
    common_cur_al_list = list(set(list_1) & set(list_2))

    print("наши current alarms:", common_cur_al_list)
    print("количество наших current alarms ", len(common_cur_al_list))
    return common_cur_al_list






#######################################################
# Main Area

list_1 = read_our_book(path=our_book_path)  # значение функции считывающей наши БС

list_2 = read_ne_list(path=ne_list_path)  # значение функции считывающей NE_list

list_5 = common_ne(list_1=list_1,list_2=list_2)   # количество наших станций NE

with open(f"alarms/NE_REP_list.json","w") as file:
    json.dump(list_5,file, indent=4, ensure_ascii=False)

list_3 = read_current_alarms(path=current_alarms_path)   # значение current alarms

list_4 = common_current_alarms(list_1=list_1, list_2=list_3)  # наши current alarms



# print(list_4)







